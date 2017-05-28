# *.xlsx laden und lesen
from openpyxl import load_workbook
# *.xlsx schreiben
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
# string-escape für sql
import pg 
pg=pg.connect(dbname='election', user='julez', passwd='lausbu3b')
# hashtags finden und schneiden
import re 
regex = re.compile('[^a-zA-Z]') 

# workbook mit unbereinigten Daten laden
wb = load_workbook('american-election-tweets.xlsx')
ws = wb['american-election-tweets']
wslen = ws.max_row-1

# wb für bereinigte Daten zum Schreiben erstellen
wb_new = Workbook()
dest_filename = 'american-election-repair.xlsx'
ws_new = wb_new.active
ws_new.title = 'bereinigt'

# Hilfsfunktion ist_enthalten für handle/ tag
def ist_enthalten (cell_value, liste):
	lenListe = len(liste)
	for name in range(0,lenListe):
		if cell_value==liste[name]:
			return True
		else:
			continue
		return False


# für Relation handle:
handle_list = []
zaehlerID = 0
ws_new['A1'] = 'handle_id'
ws_new['B1'] = 'handle'
for row in range(0, wslen):
	cell_handle_read = "{col}{row}".format(col='A', row=(row+2))
	cell_handle_id_write = "{col}{row}".format(col='A', row=(row+2))
	cell_handle_write = "{col}{row}".format(col='B', row=(row+2))
	cell_read = ws[cell_handle_read].value
	if zaehlerID==0:
		handle_list= [cell_read]
		ws_new[cell_handle_id_write] = row
		ws_new[cell_handle_write] = cell_read
		zaehlerID+=1
	elif ist_enthalten(cell_read, handle_list)==True:
		for id in range(0,len(handle_list)):
			if (handle_list[id] == cell_read):
				ws_new[cell_handle_id_write] = id
				ws_new[cell_handle_write] = cell_read
	else:
		handle_list.append(cell_read)
		ws_new[cell_handle_id_write] = zaehlerID
		ws_new[cell_handle_write] = cell_read
		zaehlerID+=1

# für Relation hashtag:
hashtag_list = []
zaehlerID = 0
ws_new['C1'] = 'text'
ws_new['D1'] = 'hashtag_id'
ws_new['E1'] = 'hashtag'
for row in range(0, wslen):
	current_hash_list = ''
	hashtag_id_list = ''
	temp=[]
	cell_text_read = "{col}{row}".format(col='B', row=(row+2))
	cell_text_write = "{col}{row}".format(col='C', row=(row+2))
	cell_hash_id_write = "{col}{row}".format(col='D', row=(row+2))
	cell_hash_write = "{col}{row}".format(col='E', row=(row+2))
	tags = str(ws[cell_text_read].value)
	cell_read = str(ws[cell_text_read].value)
	
	# text in wb_new schreiben
	text = pg.escape_string(cell_read)
	ws_new[cell_text_write] = text
	
	# zelleninhalt teilen und #finden, wie auch entfernen
	temp=[tag.strip("#") for tag in tags.split() if tag.startswith("#")]
	if (len(temp)>0):
		temp[0] = regex.sub('', temp[0])
		if (zaehlerID==0):
			temp[0] = regex.sub('', temp[0])
			hashtag_list = [temp[zaehlerID]]
			current_hash_list += str(temp[zaehlerID])
			hashtag_id_list += str(zaehlerID)
			zaehlerID+=1
		elif (zaehlerID>0):
			if ist_enthalten(temp[0], hashtag_list)==True:
				for id in range(0,len(hashtag_list)):
					if (hashtag_list[id] == temp[0]):
						hashtag_id_list += str(id)
						current_hash_list += str(temp[0])
			else:
				hashtag_list.append(temp[0])
				hashtag_id_list += str(zaehlerID)
				current_hash_list += str(temp[0])
				zaehlerID+=1
	if (len(temp)>1):
		for i in range(1,len(temp)):
			temp[i] = regex.sub('', temp[i])
			if ist_enthalten(temp[i], hashtag_list)==True:
				for id in range(0,len(hashtag_list)):
					if (hashtag_list[id] == temp[i]):
						hashtag_id_list += ' ' + str(id)
						current_hash_list += ' ' + str(temp[i])
			else:
				hashtag_list.append(temp[i])
				hashtag_id_list += ' ' + str(zaehlerID)
				current_hash_list += ' ' + str(temp[i])
				zaehlerID+=1
	else:
		hashtag_id_list += str(None)
		current_hash_list += str(None)
	ws_new[cell_hash_id_write] = hashtag_id_list
	ws_new[cell_hash_write] = current_hash_list
	
	
# für Relation tweet:
col_tweet = ['E','I','H','D'] # Auswahl der Spalten
# timestamp
tweet_list = []
zaehlerID = 0
ws_new['F1'] = 'timeTweeted'
ws_new['G1'] = 'favourite_count'
ws_new['H1'] = 'retweet_count'
ws_new['I1'] = 'original_author'
for row in range(0, wslen):
	current_hash_list = ''
	hashtag_id_list = ''
	temp=[]
	cell_time_read = "{col}{row}".format(col='E', row=(row+2))
	cell_time_write = "{col}{row}".format(col='F', row=(row+2))
	time = ws[cell_time_read].value.replace('T', ' ')
	ws_new[cell_time_write] = time		
	
	cell_fav_read = "{col}{row}".format(col='I', row=(row+2))
	cell_fav_write = "{col}{row}".format(col='G', row=(row+2))
	ws_new[cell_fav_write] = ws[cell_fav_read].value
	
	cell_retw_read = "{col}{row}".format(col='H', row=(row+2))
	cell_retw_write = "{col}{row}".format(col='H', row=(row+2))
	ws_new[cell_retw_write] = ws[cell_retw_read].value
	
	cell_auth_read = "{col}{row}".format(col='D', row=(row+2))
	cell_auth_write = "{col}{row}".format(col='I', row=(row+2))
	kl=ws[cell_auth_read].value
	auth = pg.escape_string(str(ws[cell_auth_read].value))
	ws_new[cell_auth_write] = auth
	











wb_new.save(filename = dest_filename)
